import os, json, shutil
from openpyxl import load_workbook

BASE_PDF_DIR = os.path.abspath("documents")
TEMPLATE_EXCEL_PATH = os.path.abspath("documents/grade_template.xlsx")
class InvalidPath(Exception):
    pass

class NotFound(Exception):
    pass

def process_rubric():
    with open('rubric.json', 'r') as f:
        RUBRIC = json.load(f)
    RUBRIC_KEYS = list(RUBRIC.keys())
    return {
        "rubric": RUBRIC,
        "keys": RUBRIC_KEYS
    }

def safe_path(subpath):
    full_path = os.path.abspath(os.path.join(BASE_PDF_DIR, subpath))
    if not full_path.startswith(BASE_PDF_DIR):
        raise InvalidPath("Invalid path access")
    return full_path


def list_directory(subpath):
    current_dir = safe_path(subpath)

    if not os.path.isdir(current_dir):
        raise NotFound("Directory not found")

    folders, pdfs = [], []

    for item in os.listdir(current_dir):
        item_path = os.path.join(current_dir, item)
        if os.path.isdir(item_path):
            folders.append(item)
        elif item.lower().endswith(".pdf"):
            pdfs.append(item)

    return {
        "folders": sorted(folders),
        "files": sorted(pdfs),
    }

def resolve_pdf(subpath, filename):
    full_path = safe_path(os.path.join(subpath, filename))

    if not full_path.lower().endswith(".pdf"):
        raise NotFound("Not a PDF file")

    if not os.path.exists(full_path):
        raise NotFound("PDF not found")

    directory, file = os.path.split(full_path)
    return directory, file

def parse_student_info(file_name: str):
    base_name = os.path.basename(file_name)
    name_part = os.path.splitext(base_name)[0]
    parts = name_part.split('_')
    if len(parts) < 2:
        raise ValueError("Filename must be in format LM12345_JohnDoe_Project.pdf")
    student_id = parts[0]
    student_name = ' '.join(parts[1:-1])
    return student_id, student_name

def save_student_grade(file_path: str, rubric_grades: dict, output_file="grades/all_grades.json"):
    
    print("Saving grades for:", file_path)
    # Ensure directory exists
    os.makedirs(os.path.dirname(output_file), exist_ok=True)

    student_id, student_name = parse_student_info(file_path)

    total_marks = sum(item.get("marks_awarded", 0) for item in rubric_grades.values())

    overall_comment = ". ".join(item.get("comment", "") for item in rubric_grades.values()).strip()
    # Build the student grade entry
    student_data = {
        "student_id": student_id,
        "student_name": student_name,
        "file_path": file_path,
        "total_marks": total_marks,
        "rubric": rubric_grades,
        "overall_comment": overall_comment
    }

    # Load existing JSON array if it exists
    if os.path.exists(output_file):
        try:
            with open(output_file, "r") as f:
                all_students = json.load(f)
                if not isinstance(all_students, list):
                    all_students = []
        except json.JSONDecodeError:
            all_students = []
    else:
        all_students = []

    # Check if student already exists in array (by ID)
    exists = False
    for i, s in enumerate(all_students):
        if s["student_id"] == student_id:
            all_students[i] = student_data  # update existing student
            exists = True
            break

    if not exists:
        all_students.append(student_data)  # add new student

    # Save back to JSON
    with open(output_file, "w") as f:
        json.dump(all_students, f, indent=4)
    
  
    write_grades_to_excel(
       student_data
    )

    return student_data

def create_student_excel(pdf_path):
    """
    Copies the Excel template, renames it, and saves it in the same folder as the PDF
    """

    full_pdf_path = safe_path(pdf_path)
    pdf_dir = os.path.dirname(full_pdf_path)
    os.makedirs(pdf_dir, exist_ok=True)

    student_id, student_name = parse_student_info(pdf_path)
    new_filename = f"{student_id} {student_name}.xlsx"
    new_file_path = os.path.join(pdf_dir, new_filename)

    shutil.copy(TEMPLATE_EXCEL_PATH, new_file_path)

    return new_file_path

def write_grades_to_excel(student_data):
    """
    Writes awarded marks into their respective Excel cells
    """
    excel_path = create_student_excel(student_data["file_path"])
    student_grades = student_data["rubric"]
    rubric = process_rubric()["rubric"]
    sheet_name = "Grading Sheet"
  
    wb = load_workbook(excel_path)

    
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f'Worksheet "{sheet_name}" not found. '
            'Please ensure the template has a "Grading Sheet" worksheet.'
        )
        
    sheet = wb[sheet_name]
    sheet["B2"].value = student_data["student_name"]
    sheet["B3"].value = student_data["student_id"]
    for key, grade in student_grades.items():
        if key not in rubric:
            #TODO need to handle this
            continue
        print(f"Writing {grade['marks_awarded']} to cell {rubric[key]['excelCell']} for key {key}")
        sheet[rubric[key]["excelCell"]].value = grade["marks_awarded"]
    
    result_sheet = wb.worksheets[1]
    result_sheet["A20"].value = student_data["overall_comment"]
    wb.save(excel_path)
    wb.close()